from flask import Flask, render_template, session, redirect, url_for, request
import openpyxl, os

from ai import ai_bp
from network import network_bp
from predict import predict_bp
from project import project_bp
#from plant import plant_bp  # <-- new plant detection blueprint

app = Flask(
    __name__,
    template_folder="templates",
    static_folder="static"
)
app.secret_key = "farmtech_secret"

EXCEL_FILE = "users.xlsx"

# Register blueprints
app.register_blueprint(ai_bp)
app.register_blueprint(network_bp)
app.register_blueprint(predict_bp)
app.register_blueprint(project_bp)

# Setup Excel file
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Name", "Email", "Password"])
    wb.save(EXCEL_FILE)

@app.route("/")
def home():
    user = session.get("user")
    return render_template("index.html", user=user)

@app.route("/signup", methods=["POST"])
def signup():
    name = request.form["name"]
    email = request.form["email"]
    password = request.form["password"]

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == email:
            return "Email already registered. <a href='/'>Try again</a>"

    ws.append([name, email, password])
    wb.save(EXCEL_FILE)

    session["user"] = {"name": name, "email": email}
    return redirect(url_for("home"))

@app.route("/login", methods=["POST"])
def login():
    email = request.form["email"]
    password = request.form["password"]

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1] == email and row[2] == password:
            session["user"] = {"name": row[0], "email": row[1]}
            return redirect(url_for("home"))

    return "Invalid email or password. <a href='/'>Try again</a>"

@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("home"))

if __name__ == "__main__":
    app.run(port=5000, debug=True)
